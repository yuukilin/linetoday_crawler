import os
import re
import json
import uuid
import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from google.oauth2 import service_account
from googleapiclient.discovery import build
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.text import MIMEText
import base64

MAIL_TO = os.getenv('MAIL_TO', 'yuukilin22@gmail.com')

GCP_JSON = os.getenv('GCP_JSON')
STEP = 20
BATCH = 10
HOURS_WINDOW = 12
WANT_FIN = 10
WANT_ALL = 35
MAX_TOPIC_INIT = 5
MAX_TOPIC_RELAX = 10

# Load stock names from Code.gs to keep file small
STOCK_LIST = []
_match = re.search(r"var TW_STOCK_NAMES=\[(.*?)\];", open('Code.gs').read(), re.S)
if _match:
    STOCK_LIST = json.loads('[' + _match.group(1).replace('\n','') + ']')


def http_get(url, headers):
    r = requests.get(url, headers=headers, timeout=30)
    r.raise_for_status()
    return r


def http_post_raw(url, payload, headers):
    r = requests.post(url, json=payload, headers=headers, timeout=30)
    r.raise_for_status()
    return r


def env(key):
    if key == 'MAIL_TO' and not os.getenv(key):
        return MAIL_TO
    return os.getenv(key, '')


def crawl_tab(tab, cutoff):
    root = 'https://today.line.me'
    hdr = {'User-Agent': 'Mozilla/5.0'}
    html = http_get(f'{root}/tw/v3/tab/{tab}', hdr).text
    fb = json.loads(re.search(r'<script id="__NEXT_DATA__"[^>]*>([\s\S]*?)</script>', html).group(1))['props']['pageProps']['fallback']
    lids, out = {}, []
    for v in fb.values():

    def to_art(it):
        it = it.get('article', it)
        slug = it.get('url',{}).get('hash') or it.get('canonicalUrl','').split('/')[-1]
        if not slug or slug.isdigit():
            return None
        return {
            'title': (it.get('title') or it.get('headline','')).strip(),
            'source': (it.get('publisherName') or it.get('publisher') or it.get('provider') or it.get('sourceName','')).strip(),
            'url': f'https://today.line.me/tw/v2/article/{slug}',
            'ts': it.get('publishTimeUnix',0),
            'tab': tab
        }
    for lid in lids:
        off = 0
        while True:
            api = f'{root}/api/v6/listings/{lid}?country=tw&offset={off}&length={STEP}'
            its = json.loads(http_get(api, hdr).text).get('items', [])
            if not its:
                break
            older = False
            for it in its:
                a = to_art(it)
                if not a:
                    continue
                if a['ts'] < cutoff:
                    older = True
                else:
                    out.append(a)
            if older:
                break
            off += STEP
    seen, uniq = set(), []
    for a in out:
        if a['url'] not in seen:
            seen.add(a['url'])
            uniq.append(a)
    return uniq


def drop_stock(arr):
    faang = re.compile(r'(APPLE|AAPL|MICROSOFT|MSFT|GOOGL|AMAZON|META|NFLX|NVIDIA|NVDA)', re.I)
    tw = re.compile(r'\(?\d{4}\)?')
    kw = re.compile(r'個股|ETF|股價|成交|漲停|跌停|急拉|盤中|盤後|漲幅|跌幅|%|張')
    name_re = re.compile('|'.join(STOCK_LIST)) if STOCK_LIST else re.compile(r'')
    res = []
    for a in arr:
        t = a['title']
        if not ((tw.search(t) or kw.search(t) or name_re.search(t)) and not faang.search(t)):
            res.append(a)
    return res

TOPICS = ['總體經濟','區域政治','美國科技','債券','匯率','美國政策','美歐日股市要聞','黃金','石油','新興亞洲要聞']
CODE = re.compile(r'```(?:json)?\s*|```')

def classify_batch(titles):
    salt = uuid.uuid4().hex
    prompt = f'salt:{salt}\n以下列出新聞標題（**僅根據標題判斷，勿閱讀內文**），請輸出 JSON 陣列:{{"idx":1,"yes":1,"topic":"總體經濟"}}。\n題材:{",".join(TOPICS)}\n標題:\n'
    for i, t in enumerate(titles,1):
        prompt += f'{i}. {t}\n'

        txt = r.json()['choices'][0]['message']['content']
        txt = CODE.sub('', txt).strip()
        arr = json.loads(txt)
    except Exception:
        return [None]*len(titles)
    out = [None]*len(titles)
    for o in arr:
        if o.get('yes') and 1 <= o.get('idx',0) <= len(titles):
            out[o['idx']-1] = o.get('topic') or TOPICS[0]
    return out

ALLOC = {
    'O': {'t':['總體經濟','債券','匯率'], 'color':'#FFC7CE'},
    'K': {'t':['新興亞洲要聞','美國科技'], 'color':'#C6EFCE'},
    'T': {'t':['美國政策','美歐日股市要聞'], 'color':'#FFEB9C'},
    'Y': {'t':['區域政治','黃金','石油'], 'color':'#BDD7EE'}
}

def alloc(t):
    for k,v in ALLOC.items():
        if t in v['t']:
            return k
    return ''


def sheet_blob(rows, header, filename, color_col=None):
    wb = Workbook(); ws = wb.active
    ws.append(header)
    for r in rows:
        ws.append(r)
    if color_col is not None:
        for i,r in enumerate(rows, start=2):
            key = r[color_col]
            color = ALLOC.get(key,{}).get('color','FFFFFF')
            ws.cell(i,1).fill = PatternFill(start_color=color[1:], fill_type='solid')
    wb.save(filename)
    return filename


def send_mail(files):
    if not GCP_JSON:
        return
    creds = service_account.Credentials.from_service_account_info(json.loads(GCP_JSON), scopes=['https://www.googleapis.com/auth/gmail.send'])
    service = build('gmail','v1',credentials=creds)
    message = MIMEMultipart()
    message['To'] = env('MAIL_TO')
    message['Subject'] = 'LINE Today 精選新聞（含原始與精選）'

    for f in files:
        with open(f,'rb') as fp:
            part = MIMEApplication(fp.read(), Name=os.path.basename(f))
        part['Content-Disposition'] = f'attachment; filename="{os.path.basename(f)}"'
        message.attach(part)
    raw = base64.urlsafe_b64encode(message.as_bytes()).decode()
    service.users().messages().send(userId='me', body={'raw': raw}).execute()


def fetch_and_mail():
    cutoff = int(datetime.utcnow().timestamp() - HOURS_WINDOW*3600)
    finance_raw = drop_stock(crawl_tab('finance', cutoff))
    global_raw = drop_stock(crawl_tab('global', cutoff))
    raw_seen, raw_rows = set(), []
    for a in finance_raw + global_raw:
        if a['url'] not in raw_seen:
            raw_seen.add(a['url'])
            raw_rows.append([a['tab'], a['title'], a['source'], datetime.fromtimestamp(a['ts']), a['url']])
    pending = sorted(finance_raw + global_raw, key=lambda x:x['ts'], reverse=True)
    topic_cnt, du = {}, set()
    fin_sel, other_sel = [], []
    idx, max_per = 0, MAX_TOPIC_INIT
    def is_dup(a):
        k = a.get('url') or f"{a['title']}|{a['source']}"
        if k in du:
            return True
        du.add(k)
        return False
    while (len(fin_sel) < WANT_FIN or len(fin_sel)+len(other_sel) < WANT_ALL) and idx < len(pending):
        batch = pending[idx:idx+BATCH]
        topics = classify_batch([x['title'] for x in batch])
        for a,t in zip(batch, topics):
            if not t or topic_cnt.get(t,0) >= max_per or is_dup(a):
                continue
            a['topic'] = t
            if a['tab']=='finance' and len(fin_sel) < WANT_FIN:
                fin_sel.append(a); topic_cnt[t] = topic_cnt.get(t,0)+1
            elif len(fin_sel)+len(other_sel) < WANT_ALL:
                other_sel.append(a); topic_cnt[t] = topic_cnt.get(t,0)+1
            if len(fin_sel) >= WANT_FIN and len(fin_sel)+len(other_sel) >= WANT_ALL:
                break
        idx += BATCH
        if idx>=len(pending) and len(fin_sel)+len(other_sel)<WANT_ALL and max_per<MAX_TOPIC_RELAX:
            max_per=MAX_TOPIC_RELAX; idx=0
    filt_rows = sorted(fin_sel+other_sel, key=lambda x:x['ts'], reverse=True)
    filt_rows = [[alloc(a['topic']), f"{a['title']} | {a['source']}", a['url'], a['topic']] for a in filt_rows]
    raw_file = sheet_blob(raw_rows, ['tab','title','source','datetime','url'], 'news_raw.xlsx')
    filt_file = sheet_blob(filt_rows, ['分配','標題 | 來源','連結','分類'], 'news_filtered.xlsx', 0)
    send_mail([raw_file, filt_file])

if __name__ == '__main__':
    fetch_and_mail()
